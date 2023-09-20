# -*- coding: utf-8 -*-
"""
@Time : 2022/10/3
@Auth : song
@File : FileSystem.py
@IDE  : PyCharm
@Edition: 003
@Describe: 20221002 upload: File/folder related operations v1 (judge whether the file/folder exists, whether it can access read and write permissions, enumerate the files/folders under the folder, the number of files/folders, customize the file/folder deletion operation for the specified folder, delete the specified folder
           20221003 update: 1) For the deletion of files/folders, add and delete files from the specified folder; 2) Add new files/folders; 3) Add fuzzy query file; 4) Add and rename files/folders; 5) Add Copy Paste File/Folder
           20221007 update: Convert .xlsx file to .tsv file
"""
import os
import shutil
import datetime
import platform
import traceback
import pandas as pd
import openpyxl as op


# Note that all are encapsulated as static functions
@staticmethod
def folder_is_exists(src: str):
    """
    Determine whether the folder exists
    @param src: Folder Path
    @return: Return folder path if the folder exists; Otherwise, create a folder and return the path
    """
    try:
        if not os.path.exists(src):  # Folder does not exist, create one
            os.makedirs(src)
        return src
    except (FileNotFoundError, PermissionError):
        traceback.print_exc()


@staticmethod
def file_is_exists(src: str):
    """
    Judge whether the file exists
    @param src: File Path
    @return: Return judgment results
    """
    try:
        # return os.path.isfile(src)
        return os.access(src, os.F_OK)  # access() test whether the specified file of the call exists
    except PermissionError:
        traceback.print_exc()


@staticmethod
def file_is_readable(file_absolute_path):
    """
    Judge the read permission of the file/folder
    @param file_absolute_path: The absolute path of the file
    @return: The file is readable and returns true; Otherwise, return false
    """
    try:
        if not os.path.exists(file_absolute_path):  # To determine that the file with read/write permission does not exist, an exception FileNotFoundError is thrown
            raise FileNotFoundError("file not found!")
        r_acc = os.access(file_absolute_path, os.R_OK)  # access() test whether the calling user has the specified access to the path
    except Exception as e:
        return e
    else:
        return r_acc


@staticmethod
def file_is_writable(file_absolute_path):
    """
    Judge the write permission of the file/folder
    @param file_absolute_path: The absolute path of the file
    @return: The file is writable and returns true; Otherwise, return false
    """
    try:
        if not os.path.exists(file_absolute_path):  # To determine that the file with read/write permission does not exist, an exception FileNotFoundError is thrown
            raise FileNotFoundError("file not found!")
        w_acc = os.access(file_absolute_path, os.W_OK)  # access() test whether the calling user has the specified access to the path
    except Exception as e:
        return e
    else:
        return w_acc


@staticmethod
def create_file(folder_path: str, file_name: str):
    """
    new file
    @param folder_path: The path of the folder where the file to be created is located
    @param file_name: file name
    @return: The absolute path of the file
    """
    try:
        if not os.path.exists(folder_path):  # The folder of the file to be created does not exist. Create one
            os.makedirs(folder_path)
        file_absolute_path = os.path.join(folder_path, file_name)  # os.path.join() Used to splice file paths, such as folder path+file name under the folder
        with open(file_absolute_path, 'w', encoding='utf-8') as f:
            # with open() Used to create a temporary running environment. It is no longer necessary to access files and automatically close them. After the code in the running environment is executed, it automatically exits the environment safely
            return file_absolute_path
    except (FileNotFoundError, PermissionError):
        traceback.print_exc()


@staticmethod
def create_folder(folder_path: str):
    """
    New Folder
    @param folder_path: The path of the folder to create
    @return: The absolute path of the file
    """
    try:
        if not os.path.exists(folder_path):  # The folder you want to create does not exist. Create one
            os.makedirs(folder_path)
        return folder_path
    except (FileNotFoundError, PermissionError):
        traceback.print_exc()


@staticmethod
def dirs_list(src: str):
    """
    Enumerate all folders under a folder
    @param src: Folder path to enumerate
    @return: Returns the absolute path of a group of folders in the form of a list
    """
    try:
        dirs_set = []  # Directory List
        # Traverse the folder
        for item in os.listdir(src):  # listdir() Returns the list of files or folder names contained in the specified folder
            path = os.path.join(src, item)
            if os.path.isdir(path):  # isFile() Test whether the path inside the function is a directory
                dirs_set.append(path)  # If yes, add to the end of the directory list
    except (FileNotFoundError, PermissionError):
        traceback.print_exc()
    else:
        return dirs_set


@staticmethod
def files_list(src: str):
    """
    Enumerate all files under the folder
    @param src: Folder Path
    @return: Returns the absolute path of a group of files in the form of a list
    """
    try:
        files_set = []  # File List
        files_count = 0
        # Traverse the folder
        for item in os.listdir(src):  # listdir() Returns the list of files or folder names contained in the specified folder
            path = os.path.join(src, item)  # os.path.join() Used to splice file paths, such as folder path+file name under the folder
            if os.path.isfile(path):  # isFile() Test whether the path inside the function is a standard file
                files_set.append(path)  # If yes, add to the end of the file list
    except (FileNotFoundError, PermissionError):
        traceback.print_exc()
    else:
        return files_set


@staticmethod
def dirs_count(src: str):
    """
    Number of folders under the test folder
    @param src: Folder path to test
    @return: Number of folders
    """
    try:
        dirs_num = 0
        # Traverse the folder
        for root, dirs, files in os.walk(src):  # walk() Traverse the specified folder
            # For each layer of traversal, root saves the absolute path of the currently traversed folder, dirs saves the names of all subfolders under the current folder, and files saves the names of all files under the current folder
            for each in dirs:
                dirs_num += 1
    except (FileNotFoundError, PermissionError):
        traceback.print_exc()
    else:
        return dirs_num


@staticmethod
def files_count(src: str):
    """
    Number of files under the test folder
    @param src: Folder path to test
    @return: Number of documents
    """
    try:
        files_num = 0
        # Traverse the folder
        for root, dirs, files in os.walk(src):  # walk() Traverse the specified folder
            # For each layer of traversal, root saves the absolute path of the currently traversed folder, dirs saves the names of all subfolders under the current folder, and files saves the names of all files under the current folder
            for each in files:
                files_num += 1
    except (FileNotFoundError, PermissionError):
        traceback.print_exc()
    else:
        return files_num


@staticmethod
def del_dir(src: str):
    """
    Delete the specified folder
    @param src: Current folder path
    @return: Return folder path
    """
    try:
        shutil.rmtree(src)  # Delete a folder and all its files
    except (FileNotFoundError, PermissionError):
        traceback.print_exc()
    else:
        return src


@staticmethod
def del_file(folder_path, file_name):
    """
    Delete files in the specified folder
    @param folder_path: Specify folder path
    @param file_name: file name
    @return: The absolute path of the file to delete
    """
    try:
        file_path = os.path.join(folder_path, file_name)
        os.remove(file_path)  # Delete a folder and all its files
    except (FileNotFoundError, PermissionError):
        traceback.print_exc()
    else:
        return file_path


@staticmethod
def del_files(src: str):
    """
    Delete files/folder customization under the current folder
    @param src: Current folder path
    @return: Return folder path
    """
    try:
        flag = 1
        while flag:
            current_path = os.chdir(src)  # Change the working directory to the current folder
            print("Current path\n", os.getcwd())  # Ensure current working directory
            print("List files/folders\n", os.listdir(os.getcwd()))  # Display the file list under the current folder
            obj = input("Enter the file/folder to delete\n")
            if os.path.isfile(os.path.join(src, obj)):  # User input is a file
                os.remove(obj)  # Delete File
                flag = 0
                print("List files/folders\n", os.listdir(os.getcwd()))
            elif os.path.isdir(os.path.join(src, obj)):  # User input is a folder
                shutil.rmtree(obj)  # Delete a folder and all its files
                flag = 0
                print("List files/folders\n", os.listdir(os.getcwd()))
            else:
                print("Input error, please re-enter\n")
    except (FileNotFoundError, PermissionError):
        traceback.print_exc()
    else:
        return src


sc_results_byName = []
def search_byName(path, file_name):
    """
    Search the file by the keyword of the file name in the specified path (Note: for the intermediate products searched by the file name, call the search_results_byName (path, file_name) method when using the search function
    Query by keyword of file name, namely fuzzy search
    @param path: Specify the path
    @param file_name: Keyword of the file name to find
    @return:
    """
    try:
        for each in os.listdir(path):  # Traverse the files in the given path
            file_path = os.path.join(path, each)  # os.path.join() Used to splice file paths, such as folder path+file name under the folder
            if os.path.isdir(file_path):  # The path is a folder, and search is called recursively_ File() method
                search_byName(file_path, file_name)
            elif os.path.isfile(file_path):  # The path is a file. Compare the file to be found with the file name of the file under the path
                if file_name in each:
                    global sc_results_byName  # Function internal declaration search_ Results is a global variable
                    sc_results_byName.append(file_path)  # Find the files that meet the requirements and add them to the end of the search list (using the append feature
    except (FileNotFoundError, PermissionError, TypeError):
        traceback.print_exc()
def search_results_byName(path, file_name):
    """
    Used to return search_ The search list of the file() method is the search result
    @param path:
    @param file_name:
    @return: Return search results as a list
    """
    try:
        search_byName(path, file_name)
        return sc_results_byName
    except (FileNotFoundError, PermissionError, TypeError):
        traceback.print_exc()


# @staticmethod
def file_rename(folder_path: str, old_filename: str, new_filename: str):
    """
    Rename the file/folder under the specified file directory
    @param folder_path: Specify file directory
    @param old_filename: File/folder name to rename
    @param new_filename: New file/folder name
    @return: Return the absolute path of the new file/folder
    """
    try:
        current_folder_path = os.chdir(folder_path)  # Change the working directory to the current folder
        rename_result = os.rename(old_filename, new_filename)  # Rename File/Folder
        return os.path.join(folder_path, new_filename)  # Return the absolute path of the new file/folder
    except (FileNotFoundError, FileExistsError, PermissionError, TypeError):
        traceback.print_exc()


@staticmethod
def file_copy_move(old_folder_path: str, old_file_name: str, new_folder_path: str):
    """
    Copy old files from the specified folder and move (or paste) them to the new folder
    @param old_folder_path: The folder path of the file to be copied and pasted
    @param old_file_name: File name of the file to copy and paste
    @param new_folder_path: The absolute path of the folder where the original file is to be copied and pasted
    @return: The absolute path of the newly copied file
    """
    try:
        current_path = os.chdir(old_folder_path)
        # print(os.listdir(os.getcwd()))
        tmp_file_name = "tmp"+os.path.splitext(old_file_name)[1]  # The suffix of splicing "tmp" and intercepting the original file is called intermediate product
        shutil.copyfile(old_file_name, tmp_file_name)  # Copy the original file to the current file directory
        shutil.move(tmp_file_name, new_folder_path)  # Move the intermediate products to the new file directory
        new_file_name = file_rename(new_folder_path, tmp_file_name, old_file_name)  # Rename the intermediate product to the original file name
        # current_path = os.chdir(new_folder_path)
        # print(os.listdir(os.getcwd()))
        return os.path.join(new_folder_path, old_file_name)  # Returns the absolute path of the newly copied file
    except (NotADirectoryError, FileNotFoundError, PermissionError, TypeError):
        traceback.print_exc()


@staticmethod
def folder_copy_move(old_folder_path: str, old_folder_name: str, new_folder_path: str):
    """
    Copy the old folder from the specified folder and move (or paste) it to the new folder
    @param old_folder_path: The folder path of the file to be copied and pasted
    @param old_folder_name: The name of the folder to copy and paste
    @param new_folder_path: The absolute path of the folder where the original folder is to be copied and pasted
    @return: The absolute path of the new folder
    """
    try:
        current_path = os.chdir(old_folder_path)
        # print(os.listdir(os.getcwd()))
        tmp_folder_name = "tmp"
        shutil.copytree(old_folder_name, tmp_folder_name)  # Copy the original folder to the current file directory and name it "tmp"
        shutil.move(tmp_folder_name, new_folder_path)  # Move the intermediate products to the new file directory
        new_file_name = file_rename(new_folder_path, tmp_folder_name, old_folder_name)  # Rename the intermediate product to the original file name
        # current_path = os.chdir(new_folder_path)
        # print(os.listdir(os.getcwd()))
        return os.path.join(new_folder_path, old_folder_name)  # Returns the absolute path of the newly copied folder
    except (NotADirectoryError, FileNotFoundError, PermissionError, TypeError):
        traceback.print_exc()


def xlsx_to_tsv(xlsx_path, tsv_path, sheet_num):
    """
    Convert ".xlsx" file to ".tsv" file
    @param xlsx_path: ".xlsx" file path(for example, 'C:/Users/15643/Desktop/P1-07-Drug-TargetMapping.xlsx'
    @param tsv_path: ".tsv" file path
    @param sheet_num: Sheet No. in ".xlsx" file(for example, 'Sheet1'
    @return:  ".tsv" file path
    """
    count = []
    wb = op.load_workbook(xlsx_path)
    wn = wb[sheet_num]  # Specify the name of the current worksheet
    # (Note: the case in should be consistent with that in the worksheet, such as "Sheet1“
    # print('column:', ws.max_column)
    for i in range(0, wn.max_column):
        count.append(i)
    df = pd.read_excel(xlsx_path, sheet_name=sheet_num, usecols=count, header=None)
    df.to_csv(tsv_path, header=None, sep='\t', index=False)
    return tsv_path


@staticmethod
def folder_memory_usage(src: str):
    """
    Test folder memory usage
    @param src: Current folder path
    @return: Return folder path
    """
    pass

# Update required:
# Judge the file capacity limit under folders of different operating systems

# def platform_system():
#     if platform.system() == 'Windows':
#         return 'Windows system'
#     elif platform.system() == 'Linux':
#         return 'Linux system'
#     else:
#         return 'Other system'
